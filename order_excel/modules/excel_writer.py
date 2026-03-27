import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import io
import math
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config

_HEADER_FILL  = PatternFill("solid", start_color="D9E1F2")
_HEADER_FONT  = Font(bold=True, size=9, name="Arial")
_CELL_FONT    = Font(size=9, name="Arial")
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN         = Side(style="thin")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_COL          = {}   # built at runtime

# Excel column width unit → pixels (approximate)
_COL_UNIT_TO_PX = 7.5
# Excel points → pixels
_PT_TO_PX       = 1.333


def save_manufacturer_workbook(manufacturer, rows, photo_map, barcode_map, units_per_box_map=None):
    global _COL
    _COL = {h: i + 1 for i, h in enumerate(config.OUTPUT_HEADERS)}
    wb   = Workbook()
    ws   = wb.active
    ws.title = manufacturer[:31]
    _write_header_row(ws)

    for data_row, (_, row) in enumerate(rows.iterrows(), start=2):
        item_name = str(row.get(config.COL_ITEM_NAME, ""))
        raw_sku   = row.get(config.COL_SKU, "")
        # Keep SKU as string exactly as-is to preserve leading zeros (e.g. "0407267")
        sku = str(raw_sku).strip() if raw_sku is not None else ""
        # Only strip .0 if it came in as a float with no leading zeros
        if sku.endswith(".0") and not sku.startswith("0"):
            try:
                sku = str(int(float(sku)))
            except (ValueError, TypeError):
                pass

        qty        = row.get(config.COL_ORDER_QTY, "")
        lookup_key = sku if sku.lower() not in ("nan", "", "none") else item_name

        upb = (units_per_box_map or {}).get(lookup_key)
        _write_data_row(ws, data_row, item_name, sku, qty,
                        photo_map.get(lookup_key),
                        barcode_map.get(lookup_key),
                        upb)

    out_path = _output_path(manufacturer)
    wb.save(out_path)
    print(f"  Saved -> {out_path}")
    return out_path


def _write_header_row(ws):
    for col_idx, (header, width) in enumerate(
        zip(config.OUTPUT_HEADERS, config.OUTPUT_COL_WIDTHS), start=1
    ):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30


def _write_data_row(ws, row_num, item_name, sku, qty, photo, bc, units_per_box=None):
    # Default row height
    ws.row_dimensions[row_num].height = config.ROW_HEIGHT_PT

    # Apply base style to every cell in this row
    for col_idx in range(1, len(config.OUTPUT_HEADERS) + 1):
        cell           = ws.cell(row=row_num, column=col_idx)
        cell.font      = _CELL_FONT
        cell.alignment = _CENTER
        cell.border    = _BORDER

    # ── Text values ────────────────────────────────────────────────────
    ws.cell(row=row_num, column=_COL["ITEM NAME"]).value = ""

    # Calculate TOTAL CTNS and adjusted TOTAL QTY based on Units per Box
    try:
        qty_num = int(float(str(qty))) if qty not in (None, "", "nan") else 0
    except (ValueError, TypeError):
        qty_num = 0

    if units_per_box and qty_num > 0:
        try:
            upb = int(units_per_box)

            # Always at least 1 box if qty > 0
            total_ctns = max(1, math.ceil(qty_num / upb))

            # Adjust total qty to full boxes
            total_qty = total_ctns * upb

        except (ValueError, TypeError):
            total_ctns = ""
            total_qty = qty_num
    else:
        total_ctns = ""
        total_qty = qty_num

    ws.cell(row=row_num, column=_COL["TOTAL QTY"]).value   = total_qty if total_qty else qty
    ws.cell(row=row_num, column=_COL["PCS/CTN"]).value     = units_per_box if units_per_box else ""
    ws.cell(row=row_num, column=_COL["TOTAL CTNS"]).value  = total_ctns if total_ctns else ""

    # SKU — store as text with @ number format so Excel never strips leading zeros
    clean_sku = sku if sku.lower() not in ("nan", "", "none") else ""
    cell = ws.cell(row=row_num, column=_COL["SKU"])
    cell.value         = clean_sku
    cell.number_format = "@"   # "@" = Text format in Excel — preserves "0407267"

    # Shipping mark — use real newline, not literal \n
    shipping_text = f"AN-203\n{item_name}" if item_name not in ("nan", "", "none") else "AN-203"
    ws.cell(row=row_num, column=_COL["SHIPPING MARK"]).value = shipping_text

    # ── Images ─────────────────────────────────────────────────────────
    if photo:
        _embed_image(ws, photo, row_num, _COL["ITEM PHOTO"],
                     config.PHOTO_WIDTH_PX, config.PHOTO_HEIGHT_PX)

    if bc:
        _embed_barcode(ws, bc, row_num)


def _embed_barcode(ws, bc, row_num):
    """
    Fit barcode into the BARCODE column width.
    Scale height proportionally and expand row if needed.
    """
    col_letter  = get_column_letter(_COL["BARCODE"])
    col_width_u = ws.column_dimensions[col_letter].width or config.OUTPUT_COL_WIDTHS[_COL["BARCODE"] - 1]
    col_px      = int(col_width_u * _COL_UNIT_TO_PX)

    bc_w, bc_h  = bc.size
    scale       = col_px / bc_w
    disp_w      = col_px
    disp_h      = int(bc_h * scale)

    # Row height in points — expand if barcode is taller than default
    needed_pt   = disp_h / _PT_TO_PX
    if needed_pt > ws.row_dimensions[row_num].height:
        ws.row_dimensions[row_num].height = needed_pt

    _embed_image(ws, bc, row_num, _COL["BARCODE"], disp_w, disp_h)


def _embed_image(ws, pil_img, row, col, w, h):
    buf        = io.BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    xl_img        = XLImage(buf)
    xl_img.width  = w
    xl_img.height = h
    ws.add_image(xl_img, f"{get_column_letter(col)}{row}")


def _output_path(manufacturer):
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    safe = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in manufacturer)
    return os.path.join(config.OUTPUT_DIR, f"{safe}.xlsx")